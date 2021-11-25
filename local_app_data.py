from openpyxl import load_workbook


def save_theme(theme, primary, accent):
    wb = load_workbook('theme_data.xlsx')
    ws = wb['themes']

    ws['B2'].value = theme
    ws['B3'].value = primary
    ws['B4'].value = accent

    wb.save('theme_data.xlsx')


def load_theme():
    wb = load_workbook('theme_data.xlsx')
    ws = wb['themes']
    theme = ws['B2'].value
    return theme


def load_primary():
    wb = load_workbook('theme_data.xlsx')
    ws = wb['themes']
    primary = ws['B3'].value
    return primary


def load_accent():
    wb = load_workbook('theme_data.xlsx')
    ws = wb['themes']
    accent = ws['B4'].value
    return accent


def check_status():
    wb = load_workbook('theme_data.xlsx')
    ws = wb['profile_status']
    status = ws['B1'].value

    return status


def set_status(s, strng, strng2):
    wb = load_workbook('theme_data.xlsx')
    ws = wb['profile_status']

    if s == 1:
        ws['B1'].value = s
        ws['B2'].value = strng
        ws['B3'].value = strng2
    elif s == 0:
        ws['B1'].value = s
        ws['B2'].value = strng
        ws['B3'].value = strng2
    wb.save('theme_data.xlsx')


def get_profile(c):
    wb = load_workbook('theme_data.xlsx')
    ws = wb['profile_status']
    if c == 0:
        info = ws['B2'].value
        print(info)
        return info
    else:
        info = ws['B3'].value
        print(info)
        return info
